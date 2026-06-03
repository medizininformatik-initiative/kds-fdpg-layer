#!/usr/bin/env python3
from __future__ import annotations

"""Generate Datenkatalog (Data Dictionary) markdown pages for the FDPG KDS Obligations IG.

For each module, reads:
  - MII parent profiles from the FHIR package cache (snapshot elements with mustSupport)
  - FDPG FSH files (profile name, parent, FHIR id)
  - Existing modul-*.md pages for sub-section groupings (ICU, Onkologie)

Outputs per-module markdown into input/pagecontent/datenkatalog-{module}.md
"""

import json
import os
import re
import sys
from pathlib import Path

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

FHIR_CACHE = Path.home() / ".fhir" / "packages"
PROJECT_ROOT = Path(__file__).resolve().parent.parent
FSH_DIR = PROJECT_ROOT / "input" / "fsh" / "obligations"
PAGECONTENT_DIR = PROJECT_ROOT / "input" / "pagecontent"
LM_SUPPLEMENT_DIR = PROJECT_ROOT / "input" / "data" / "lm-supplements"

# Module configs: key = obligation folder name, value = dict with package info
# For basis, the package is "base" but the folder is "basis"
MODULES = {
    "basis": {
        "package": "de.medizininformatikinitiative.kerndatensatz.base",
        "version": "2026.0.0",
        "title": "Basisdaten",
        "subtitle": "Person, Diagnose, Prozedur, Fall",
        "simplifier_url": "https://simplifier.net/packages/de.medizininformatikinitiative.kerndatensatz.base/2026.0.0",
    },
    "labor": {
        "package": "de.medizininformatikinitiative.kerndatensatz.laborbefund",
        "version": "2026.0.1",
        "title": "Laborbefund",
        "simplifier_url": "https://simplifier.net/packages/de.medizininformatikinitiative.kerndatensatz.laborbefund/2026.0.1",
    },
    "medikation": {
        "package": "de.medizininformatikinitiative.kerndatensatz.medikation",
        "version": "2026.0.1",
        "title": "Medikation",
        "simplifier_url": "https://simplifier.net/packages/de.medizininformatikinitiative.kerndatensatz.medikation/2026.0.1",
    },
    "biobank": {
        "package": "de.medizininformatikinitiative.kerndatensatz.biobank",
        "version": "2026.0.1",
        "title": "Biobank",
        "simplifier_url": "https://simplifier.net/packages/de.medizininformatikinitiative.kerndatensatz.biobank/2026.0.1",
    },
    "studie": {
        "package": "de.medizininformatikinitiative.kerndatensatz.studie",
        "version": "2026.0.2",
        "title": "Studie",
        "simplifier_url": "https://simplifier.net/packages/de.medizininformatikinitiative.kerndatensatz.studie/2026.0.2",
    },
    "molgen": {
        "package": "de.medizininformatikinitiative.kerndatensatz.molgen",
        "version": "2026.0.4",
        "title": "Molekulargenetik",
        "simplifier_url": "https://simplifier.net/packages/de.medizininformatikinitiative.kerndatensatz.molgen/2026.0.4",
    },
    "patho": {
        "package": "de.medizininformatikinitiative.kerndatensatz.patho",
        "version": "2026.0.1",
        "title": "Pathologiebefund",
        "simplifier_url": "https://simplifier.net/packages/de.medizininformatikinitiative.kerndatensatz.patho/2026.0.1",
    },
    "icu": {
        "package": "de.medizininformatikinitiative.kerndatensatz.icu",
        "version": "2026.0.2",
        "title": "Intensivmedizin",
        "simplifier_url": "https://simplifier.net/packages/de.medizininformatikinitiative.kerndatensatz.icu/2026.0.2",
    },
    "bildgebung": {
        "package": "de.medizininformatikinitiative.kerndatensatz.bildgebung",
        "version": "2026.0.0",
        "title": "Bildgebung",
        "simplifier_url": "https://simplifier.net/packages/de.medizininformatikinitiative.kerndatensatz.bildgebung/2026.0.0",
    },
    "seltene": {
        "package": "de.medizininformatikinitiative.kerndatensatz.seltene",
        "version": "2026.0.1",
        "title": "Seltene Erkrankungen",
        "simplifier_url": "https://simplifier.net/packages/de.medizininformatikinitiative.kerndatensatz.seltene/2026.0.1",
    },
    "onkologie": {
        "package": "de.medizininformatikinitiative.kerndatensatz.onkologie",
        "version": "2026.0.3",
        "title": "Onkologie",
        "simplifier_url": "https://simplifier.net/packages/de.medizininformatikinitiative.kerndatensatz.onkologie/2026.0.3",
    },
    "consent": {
        "package": "de.medizininformatikinitiative.kerndatensatz.consent",
        "version": "2026.0.1-rc-2",
        "title": "Einwilligung",
        "simplifier_url": "https://simplifier.net/packages/de.medizininformatikinitiative.kerndatensatz.consent/2026.0.1-rc-2",
    },
    "dokument": {
        "package": "de.medizininformatikinitiative.kerndatensatz.dokument",
        "version": "2026.0.1",
        "title": "Dokument",
        "simplifier_url": "https://simplifier.net/packages/de.medizininformatikinitiative.kerndatensatz.dokument/2026.0.1",
    },
    "mtb": {
        "package": "de.medizininformatikinitiative.kerndatensatz.mtb",
        "version": "2026.0.1",
        "title": "MTB",
        "simplifier_url": "https://simplifier.net/packages/de.medizininformatikinitiative.kerndatensatz.mtb/2026.0.1",
    },
    "proms": {
        "package": "de.medizininformatikinitiative.kerndatensatz.pros",
        "version": "2026.3.0",
        "title": "PRO",
        "simplifier_url": "https://simplifier.net/packages/de.medizininformatikinitiative.kerndatensatz.pros/2026.3.0",
    },
    "mikrobiologie": {
        "package": "de.medizininformatikinitiative.kerndatensatz.mikrobiologie",
        "version": "2027.0.0-alpha.3",
        "title": "Mikrobiologie",
        "simplifier_url": "https://simplifier.net/packages/de.medizininformatikinitiative.kerndatensatz.mikrobiologie/2027.0.0-alpha.3",
    },
}

# Sub-section groupings for large modules.  Extracted from existing modul-*.md.
# Maps module_name -> list of (section_title, [parent_profile_names])
# Profiles not listed in any section go into an "Weitere Profile" catch-all.
MODULE_SECTIONS: dict[str, list[tuple[str, list[str]]]] = {}

TRANSLATION_EXT_URL = "http://hl7.org/fhir/StructureDefinition/translation"

# Obligation labels for the markdown table
OBLIGATION_LABEL = "MustSupport"

CANONICAL_LABELS_FILE = PROJECT_ROOT / "input" / "data" / "canonical-labels.json"
FIELD_CONFIG_FILE = PROJECT_ROOT / "input" / "data" / "field_config.json"


def load_field_config() -> dict:
    """Load field_config.json and return per-base-resource pre-select element IDs.

    Returns dict of base-StructureDefinition URL → set of element IDs that should
    be pre-selected (recommend.always in field_config.json).
    """
    if not FIELD_CONFIG_FILE.exists():
        return {}
    with open(FIELD_CONFIG_FILE) as f:
        cfg = json.load(f)
    result = {}
    for base_url, prof in cfg.get("profiles", {}).items():
        ids = set()
        for entry in prof.get("recommend", {}).get("always", []):
            elem_id = entry.get("pattern", {}).get("id")
            if elem_id:
                ids.add(elem_id)
        if ids:
            result[base_url] = ids
    return result


def get_pre_select_ids(profile_data: dict, field_config: dict) -> set:
    """Return the set of element IDs flagged pre-select for this profile."""
    base = profile_data.get("baseDefinition")
    if base and base in field_config:
        return field_config[base]
    return set()

# Override english-leaning leaf labels coming from upstream MII Logical Models.
# Maps original LM element name -> {de, en} replacement.
# Source: docs/datenkatalog-styleguide.md §3 / manual review of LM exports.
LM_LABEL_OVERRIDES: dict[str, dict] = {
    "FollowUpStatus":         {"de": "NachsorgeStatus",          "en": "FollowUpStatus"},
    "ResponseBeurteilung":    {"de": "AnsprechenBeurteilung",    "en": "ResponseAssessment"},
    "PatternTyp":             {"de": "MusterTyp",                "en": "PatternType"},
    "IntraoperativesImaging": {"de": "IntraoperativeBildgebung", "en": "IntraoperativeImaging"},
    "HPOExcluded":            {"de": "HPOAusgeschlossen",        "en": "HPOExcluded"},
}


def load_module_labels(module_name: str) -> dict:
    """Load module-specific labels from input/data/element-labels-<module>.json.
    Returns {} if file doesn't exist. Mirrors enrich-fsh-translations.py."""
    path = PROJECT_ROOT / "input" / "data" / f"element-labels-{module_name}.json"
    if not path.exists():
        return {}
    with open(path) as f:
        data = json.load(f)
    return {k: v for k, v in data.get("elements", {}).items() if not k.startswith("_")}


def merge_module_labels(canonical: dict, module_name: str) -> dict:
    """Return a canonical-shaped dict with module labels overlaid on the
    'elements' section. Module entries override canonical fields per-key
    but missing fields fall back to canonical (deep merge).
    """
    module = load_module_labels(module_name)
    if not module:
        return canonical
    merged = dict(canonical or {})
    merged_elements = dict((canonical or {}).get("elements", {}))
    for key, mod_entry in module.items():
        if key in merged_elements:
            merged_elements[key] = {**merged_elements[key], **mod_entry}
        else:
            merged_elements[key] = mod_entry
    merged["elements"] = merged_elements
    return merged


def load_canonical_labels() -> dict:
    """Load canonical DE+EN labels for fallback when profile/LM lacks translations.

    Returns dict with 'elements' (path → labels) and 'coding_systems' (URL → labels).
    """
    if not CANONICAL_LABELS_FILE.exists():
        return {"elements": {}, "coding_systems": {}}
    with open(CANONICAL_LABELS_FILE) as f:
        data = json.load(f)
    elements = {k: v for k, v in data.get("elements", {}).items() if not k.startswith("_")}
    systems = {k: v for k, v in data.get("coding_systems", {}).items() if not k.startswith("_")}
    return {"elements": elements, "coding_systems": systems}


def _strip_slices_except_extension(path: str) -> str:
    """Drop ':sliceName' from each segment, except segments starting with
    'extension:' (named extensions carry semantic per slice, not inheritable).
    """
    parts = []
    for seg in path.split("."):
        if seg.startswith("extension:") or ":" not in seg:
            parts.append(seg)
        else:
            parts.append(seg.split(":", 1)[0])
    return ".".join(parts)


def lookup_canonical(canonical: dict, display_path: str) -> dict:
    """Look up canonical labels for a profile element display path.

    Tries exact match first, then a slice-strip fallback that preserves
    'extension:X' slices (named extensions are semantically specific and
    must not collapse to the generic 'extension' canonical entry).
    Coding-slice harmonisation uses the coding_systems table at the call site.
    """
    elements = canonical.get("elements", {})
    if display_path in elements:
        return elements[display_path]
    base = _strip_slices_except_extension(display_path)
    if base != display_path and base in elements:
        return elements[base]
    return {}


def get_coding_system(elem: dict) -> str:
    """Extract the coding system URL from a coding-slice element.

    Looks at patternCoding.system, pattern.system, or fixedUri.
    Returns empty string if not found.
    """
    pc = elem.get("patternCoding") or {}
    if pc.get("system"):
        return pc["system"]
    p = elem.get("pattern") or {}
    if p.get("system"):
        return p["system"]
    if elem.get("fixedUri"):
        return elem["fixedUri"]
    return ""


# Fallback display names for coding-slice names (mirrors enrich-fsh-translations.py).
SLICE_NAME_ALIASES = {
    "obds":         "oBDS",
    "sct":          "SNOMED CT",
    "snomed":       "SNOMED CT",
    "snomed-ct":    "SNOMED CT",
    "loinc":        "LOINC",
    "icd10-gm":     "ICD-10-GM",
    "icd-10-gm":    "ICD-10-GM",
    "icd-o-3":      "ICD-O-3",
    "icdo3":        "ICD-O-3",
    "alpha-id":     "Alpha-ID",
    "orphanet":     "Orphanet",
    "mondo":        "MONDO",
    "omim":         "OMIM",
    "hpo":          "HPO",
    "atc":          "ATC",
    "ops":          "OPS",
    "pzn":          "Pharmazentralnummer",
    "ucum":         "UCUM",
    "iso-3166":     "ISO 3166",
    "iso-3166-2":   "ISO 3166-2",
    "ieee-11073":   "IEEE 11073",
    "v2-microbiology": "HL7 v2 Microbiology",
    "lnc":          "LOINC",
    "kdl":          "KDL",
    "xds":          "IHE XDS",
    "miabis-type":  "MIABIS Specimen Type",
    "kdsicu-category": "KDS ICU",
    "meddra":       "MedDRA",
    "section-type": "Befundabschnitt",
    "Beatmung":     "Beatmung",
    "vs-cat":       "Vital Signs",
    "consentCategory": "Consent Category",
    "resultType":   "Result Type",
}


def lookup_coding_system_labels(canonical: dict, system_url: str) -> dict:
    """Look up DE+EN labels for a coding system URL."""
    return canonical.get("coding_systems", {}).get(system_url, {})


def coding_system_display_from_slice(slice_name: str) -> dict:
    """Return DE+EN display for a slice name from SLICE_NAME_ALIASES.
    Used as a fallback when the slice's codesystem URL is non-canonical
    (e.g., MII-specific value-set CSs).
    """
    alias = SLICE_NAME_ALIASES.get(slice_name.lower())
    if not alias:
        return {}
    return {"de_short": alias, "en_short": alias}

# Logical model config: module name → LM filenames + FHIR mapping identity strings.
# Modules without an entry (or without FHIR mappings) simply get no Fachbegriff column.
LOGICAL_MODEL_CONFIG: dict[str, dict] = {
    "basis": {
        "lm_files": [
            "StructureDefinition-mii-lm-diagnose.json",
            "StructureDefinition-mii-lm-prozedur.json",
        ],
        "supplement_files": [
            "lm-supplement-person.json",
            "lm-supplement-fall.json",
        ],
        "fhir_identities": ["FHIR"],
    },
    "labor": {
        "lm_files": ["StructureDefinition-mii-lm-labor.json"],
        "fhir_identities": ["FHIR"],
    },
    "medikation": {
        "lm_files": ["StructureDefinition-mii-lm-medikation.json"],
        "fhir_identities": ["FHIR"],
    },
    "biobank": {
        "lm_files": ["StructureDefinition-Biobank.json"],
        "fhir_identities": ["FHIR"],
    },
    "molgen": {
        "lm_files": ["StructureDefinition-LogicalModelMolGen.json"],
        "fhir_identities": ["FHIR"],
    },
    "bildgebung": {
        "lm_files": ["StructureDefinition-mii-lm-bildgebung.json"],
        "fhir_identities": ["FHIR"],
    },
    "seltene": {
        "lm_files": ["StructureDefinition-mii-lm-seltene.json"],
        "fhir_identities": ["FHIR"],
    },
    "onkologie": {
        "lm_files": [
            "StructureDefinition-mii-lm-onko.json",
            "StructureDefinition-mii-lm-onko-organspezifische-zusatzmodule.json",
            "StructureDefinition-mii-lm-mvgenomseq-onkologie.json",
        ],
        "fhir_identities": ["FHIR", "MVGenomSeq-Datenkranz-to-MII-FHIR"],
    },
    "mtb": {
        "lm_files": ["StructureDefinition-mii-lm-mtb.json"],
        "fhir_identities": ["mii-map-mtb"],
    },
    "proms": {
        "lm_files": ["StructureDefinition-mii-lm-pro.json"],
        "fhir_identities": ["FHIR"],
    },
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def get_package_dir(module_cfg: dict) -> Path:
    """Return path to the FHIR package directory."""
    dirname = f"{module_cfg['package']}#{module_cfg['version']}"
    return FHIR_CACHE / dirname / "package"


def load_profile_json(package_dir: Path, profile_name: str) -> dict | None:
    """Find and load a StructureDefinition JSON by profile name from a package dir.

    Handles both standard naming (StructureDefinition-mii-pr-*.json) and the
    consent module's special naming (Profile_MII_*.json).
    """
    # Search all JSON files in the package dir for one with matching name
    for f in package_dir.glob("*.json"):
        if not (f.name.startswith("StructureDefinition-") or f.name.startswith("Profile_MII_")):
            continue
        try:
            with open(f) as fh:
                data = json.load(fh)
            if data.get("resourceType") == "StructureDefinition" and data.get("name") == profile_name:
                return data
        except (json.JSONDecodeError, KeyError):
            continue
    return None


_EN_LEAK_RE = re.compile(
    r'(?i)^(May be used|Defined by|Identifies|Indicates|A reference|This (resource|element|condition)|Code defined by|Symbol in syntax|Healthcare event|Who has|Business Identifier|Type of observation|Classification of|Identity of the terminology|Version of the system|Plain text representation|Optional Extensions|Conditions associated|Free text|Additional content|Logical id|Metadata about|The healthcare event)'
)


def _looks_english_as_de(text: str) -> bool:
    """Heuristic: text claims to be DE but reads as English-as-DE leak."""
    if not text:
        return False
    if re.search(r'[äöüÄÖÜß]', text):
        return False  # umlauts → really German
    if re.search(r'\b(der|die|das|und|für|nach|mit|von|zur|zum|bei|als|im|am|ist|sind|werden)\b', text, re.IGNORECASE):
        return False  # German function words → really German
    return bool(_EN_LEAK_RE.match(text))


def extract_translation(element: dict, field: str, lang: str) -> str:
    """Extract a translation string from _short or _definition.

    For de-DE, filters out 'English-as-DE leak' translations (upstream
    profile-team mis-tagged an English string with lang=de-DE). Callers
    then fall back to canonical / module labels which are reliably German.
    """
    underscored = f"_{field}"
    ext_container = element.get(underscored, {})
    if not ext_container:
        return ""
    for ext in ext_container.get("extension", []):
        if ext.get("url") != TRANSLATION_EXT_URL:
            continue
        found_lang = ""
        found_content = ""
        for sub in ext.get("extension", []):
            if sub.get("url") == "lang":
                found_lang = sub.get("valueCode", "")
            elif sub.get("url") == "content":
                found_content = sub.get("valueString", "")
        if found_lang == lang:
            if lang == "de-DE" and _looks_english_as_de(found_content):
                continue  # leak — let canonical/module fallback take over
            return found_content
    return ""


# Infrastructure/metadata elements excluded per Styleguide §1 and field_config.json
# default.exclude. Applied at the top-level (depth 2) of any resource.
METADATA_TOP_ELEMENTS = {
    "id", "meta", "implicitRules", "language",
    "text", "contained", "modifierExtension",
}


def get_ms_elements(profile_data: dict) -> list[dict]:
    """Extract top-level mustSupport elements from profile snapshot.

    Uses element.id (not path) to determine depth and slice membership:
    - Include elements at depth <= 2 (e.g. Observation.code)
    - Include named slices at depth 2 (e.g. Observation.identifier:analyseBefundCode)
    - Include named coding slices at depth 3 (e.g. Observation.code.coding:icd-10-gm)
    - Exclude infrastructure elements (id, meta, meta.*, text, contained, ...)
    - Exclude deeper elements (e.g. Observation.code.coding.system)
    """
    elements = profile_data.get("snapshot", {}).get("element", [])
    result = []
    for elem in elements:
        if not elem.get("mustSupport"):
            continue
        elem_id = elem.get("id", elem.get("path", ""))

        segments = elem_id.split(".")
        depth = len(segments)

        if depth >= 2:
            top = segments[1].split(":")[0]  # strip slice suffix
            if top in METADATA_TOP_ELEMENTS:
                continue

        if depth <= 2:
            result.append(elem)
        elif depth == 3 and ":" in segments[2]:
            # Named slice on a sub-element (e.g. code.coding:icd-10-gm)
            result.append(elem)

    return result


def element_display_path(element: dict, resource_type: str) -> str:
    """Return the display path for an element (strip resource type prefix).

    Uses element.id to preserve slice names, e.g.:
    - Observation.identifier:analyseBefundCode -> identifier:analyseBefundCode
    - Observation.component:SystolicBP -> component:SystolicBP
    """
    elem_id = element.get("id", element.get("path", ""))
    prefix = resource_type + "."
    if elem_id.startswith(prefix):
        return elem_id[len(prefix):]
    return elem_id


def md_escape(text: str) -> str:
    """Escape special markdown characters for table cells and collapse whitespace."""
    if not text:
        return ""
    text = text.replace("|", "\\|")
    text = text.replace("*", "\\*")
    text = re.sub(r"\s+", " ", text).strip()
    # Truncate very long text
    if len(text) > 200:
        text = text[:197] + "..."
    return text


# ---------------------------------------------------------------------------
# Logical model helpers
# ---------------------------------------------------------------------------

def load_logical_models(module_name: str, package_dir: Path) -> list[dict]:
    """Load logical model JSON files for a module from its FHIR package."""
    config = LOGICAL_MODEL_CONFIG.get(module_name)
    if not config:
        return []
    models = []
    for filename in config["lm_files"]:
        filepath = package_dir / filename
        if not filepath.exists():
            continue
        try:
            with open(filepath) as f:
                data = json.load(f)
            if data.get("kind") == "logical":
                models.append(data)
        except (json.JSONDecodeError, KeyError):
            continue
    # Also load supplement files (locally maintained FHIR mappings)
    for filename in config.get("supplement_files", []):
        filepath = LM_SUPPLEMENT_DIR / filename
        if not filepath.exists():
            continue
        try:
            with open(filepath) as f:
                data = json.load(f)
            if data.get("kind") == "logical":
                models.append(data)
        except (json.JSONDecodeError, KeyError):
            continue
    return models


def _extract_extension_url_from_mapping(fhir_map: str) -> str:
    """Extract extension URL from a FHIR mapping string."""
    fhir_map = fhir_map.split("|")[0].strip()  # handle pipe alternatives
    m = re.search(r"\.extension\('([^']+)'\)", fhir_map)
    if m:
        return m.group(1)
    m = re.search(r"\.extension\.where\(url='([^']+)'\)", fhir_map)
    if m:
        return m.group(1)
    return ""


def _get_extension_url_from_element(element: dict) -> str:
    """Extract extension URL from profile element's type[0].profile[0]."""
    types = element.get("type", [])
    if types:
        profiles = types[0].get("profile", [])
        if profiles:
            return profiles[0]
    return ""


def parse_fhir_mapping(fhir_map: str) -> tuple[str, str] | None:
    """Parse a FHIR mapping string to (resource_type, top_level_element).

    Extracts only the first path segment after the resource type, stripping
    FHIRPath functions (.where(), .extension(), .select()) and pipe alternatives.
    Returns None for unparseable strings.
    """
    if not fhir_map or not fhir_map[0].isupper():
        return None
    # Top-level extension mappings: extract URL as specific key before cleaning
    # (cleaning strips .extension(...) which loses the URL information)
    rt_ext = re.match(r"^([A-Z][a-zA-Z]+)\.extension", fhir_map)
    if rt_ext:
        ext_url = _extract_extension_url_from_mapping(fhir_map)
        if ext_url:
            return (rt_ext.group(1), f"ext:{ext_url}")
        return None  # bare extension without URL, skip
    # Strip FHIRPath function calls (preserving surrounding path)
    cleaned = re.sub(r"\.where\([^)]*\)", "", fhir_map)
    cleaned = re.sub(r"\.extension\([^)]*\)", "", cleaned)
    cleaned = re.sub(r"\.select\([^)]*\)", "", cleaned)
    # Resolve pipe alternatives: .(a|b|c) → .a
    cleaned = re.sub(r"\(([a-zA-Z\[\]x]+)\|[^)]+\)", r"\1", cleaned)
    match = re.match(r"^([A-Z][a-zA-Z]+)(?:\.([a-zA-Z\[\]x]+))?", cleaned)
    if not match:
        return None
    return (match.group(1), match.group(2) or "")


def build_lm_lookup(
    logical_models: list[dict], fhir_identities: list[str],
) -> dict[tuple[str, str], dict]:
    """Build (resource_type, element) → info lookup from logical models.

    Each entry: {"labels": [str], "short": str, "definition": str}.
    Prefers data from shallow (depth-1) FHIR mappings over deeper ones.
    The label is the leaf segment of the logical model element ID.
    """
    primary: dict[tuple[str, str], dict] = {}   # depth-1 matches
    secondary: dict[tuple[str, str], dict] = {}  # deeper matches

    for lm in logical_models:
        for elem in lm.get("snapshot", {}).get("element", []):
            for m in elem.get("mapping", []):
                if m.get("identity") not in fhir_identities:
                    continue
                fhir_map = m.get("map", "")
                parsed = parse_fhir_mapping(fhir_map)
                if not parsed or not parsed[1]:
                    continue  # skip resource-level mappings

                resource_type, top_element = parsed
                label = elem.get("id", "").split(".")[-1]
                if not label:
                    continue

                key = (resource_type, top_element)
                is_primary = bool(re.match(
                    r"^[A-Z][a-zA-Z]+\.[a-zA-Z\[\]x:]+$", fhir_map.strip(),
                ))

                target = primary if is_primary else secondary
                if key not in target:
                    target[key] = {
                        "labels": [],
                        "short": elem.get("short", ""),
                        "definition": elem.get("definition", ""),
                    }
                if label not in target[key]["labels"]:
                    target[key]["labels"].append(label)

    # Merge: prefer primary, fall back to secondary
    result: dict[tuple[str, str], dict] = {}
    for key in set(primary) | set(secondary):
        result[key] = primary.get(key, secondary.get(key, {}))
    return result


def _resolve_lm_key(
    lm_lookup: dict[tuple[str, str], dict],
    resource_type: str,
    display_path: str,
    element: dict | None = None,
) -> dict | None:
    """Find the LM lookup entry for a profile element's display path.

    Handles slice names (stripped before lookup), choice type [x] matching,
    and extension URL matching via element type profiles.
    """
    # Extension elements: match by URL from element's type[0].profile[0]
    if display_path.startswith("extension:") and element is not None:
        ext_url = _get_extension_url_from_element(element)
        if ext_url:
            entry = lm_lookup.get((resource_type, f"ext:{ext_url}"))
            if entry:
                return entry

    # Strip slice suffix: 'identifier:analyseBefundCode' → 'identifier'
    base = re.sub(r":[^.]+", "", display_path)
    # Only match top-level element (first segment)
    top = base.split(".")[0]

    key = (resource_type, top)
    entry = lm_lookup.get(key)

    # Choice type fallback: 'value[x]' → try 'valueQuantity' etc.
    if not entry and "[x]" in top:
        prefix = top.replace("[x]", "")
        for (rt, elem), e in lm_lookup.items():
            if rt == resource_type and elem.startswith(prefix) and elem != top:
                entry = e
                break

    return entry


def lookup_lm_label(
    lm_lookup: dict[tuple[str, str], dict],
    resource_type: str,
    display_path: str,
    element: dict | None = None,
    lang: str = "de",
) -> str:
    """Resolve a profile element to its logical model concept name(s).

    Applies LM_LABEL_OVERRIDES to replace english-leaning upstream LM labels
    with curated German/English equivalents. Pass lang="en" for EN labels;
    falls back to the original LM label if no override is defined.
    """
    entry = _resolve_lm_key(lm_lookup, resource_type, display_path, element)
    if not entry:
        return ""
    labels = [LM_LABEL_OVERRIDES.get(l, {}).get(lang, l) for l in entry["labels"]]
    if len(labels) <= 2:
        return ", ".join(labels)
    return ", ".join(labels[:2]) + ", ..."


def lookup_lm_description(
    lm_lookup: dict[tuple[str, str], dict],
    resource_type: str,
    display_path: str,
    element: dict | None = None,
) -> str:
    """Resolve a profile element to its logical model description.

    Prefers 'definition', falls back to 'short'.
    """
    entry = _resolve_lm_key(lm_lookup, resource_type, display_path, element)
    if not entry:
        return ""
    return entry.get("definition") or entry.get("short", "")


# ---------------------------------------------------------------------------
# Parse FSH files for profile metadata
# ---------------------------------------------------------------------------

def parse_fsh_profiles(module_name: str) -> list[dict]:
    """Parse FSH files in the module directory and return profile metadata.

    Returns list of dicts with keys: fdpg_name, parent_name, fdpg_id, description,
    element_comments (dict mapping element path to comment text).
    """
    module_dir = FSH_DIR / module_name
    if not module_dir.is_dir():
        print(f"  WARNING: FSH directory not found: {module_dir}", file=sys.stderr)
        return []

    profiles = []
    for fsh_file in sorted(module_dir.glob("*.fsh")):
        with open(fsh_file) as f:
            content = f.read()

        # Extract Profile, Parent, Id
        profile_match = re.search(r"^Profile:\s*(\S+)", content, re.MULTILINE)
        parent_match = re.search(r"^Parent:\s*(\S+)", content, re.MULTILINE)
        id_match = re.search(r"^Id:\s*(\S+)", content, re.MULTILINE)

        if not (profile_match and parent_match and id_match):
            continue

        # Extract Description (may be multi-line in quotes)
        desc_match = re.search(
            r'^Description:\s*"((?:[^"\\]|\\.)*)"\s*$',
            content, re.MULTILINE | re.DOTALL,
        )
        description = desc_match.group(1).replace('\\"', '"') if desc_match else ""

        # Extract German title translation: insert Translation(^title, de-DE, ...)
        title_de_match = re.search(
            r'insert\s+Translation\(\^title,\s*de-DE,\s*(.+?)\)',
            content,
        )
        title_de = title_de_match.group(1).strip() if title_de_match else ""

        # Extract element-level ^comment rules
        # Matches: * some.path ^comment = "..."
        element_comments: dict[str, str] = {}
        for m in re.finditer(
            r'^\*\s+([\w.\[\]:]+)\s+\^comment\s*=\s*"((?:[^"\\]|\\.)*)"',
            content, re.MULTILINE,
        ):
            elem_path = m.group(1)
            comment_text = m.group(2).replace('\\"', '"')
            element_comments[elem_path] = comment_text

        profiles.append({
            "fdpg_name": profile_match.group(1),
            "parent_name": parent_match.group(1),
            "fdpg_id": id_match.group(1),
            "description": description,
            "title_de": title_de,
            "element_comments": element_comments,
        })

    return profiles


# ---------------------------------------------------------------------------
# Parse section groupings from existing modul-*.md pages
# ---------------------------------------------------------------------------

def parse_module_sections(module_name: str) -> list[tuple[str, list[str]]]:
    """Parse existing modul-*.md to extract section groupings.

    Returns list of (section_title, [parent_profile_names]).
    For the basis module, maps person/diagnose/prozedur/fall pages.
    """
    if module_name == "basis":
        return _parse_basis_sections()

    # Map module name to page file
    page_map = {
        "labor": "modul-labor.md",
        "medikation": "modul-medikation.md",
        "biobank": "modul-biobank.md",
        "studie": "modul-studie.md",
        "molgen": "modul-molgen.md",
        "patho": "modul-patho.md",
        "icu": "modul-icu.md",
        "bildgebung": "modul-bildgebung.md",
        "seltene": "modul-seltene.md",
        "onkologie": "modul-onkologie.md",
        "consent": "modul-consent.md",
        "dokument": "modul-dokument.md",
        "mtb": "modul-mtb.md",
        "proms": "modul-pros.md",
    }

    page_filename = page_map.get(module_name, "")
    if not page_filename:
        return []
    page_file = PAGECONTENT_DIR / page_filename
    if not page_file.exists():
        return []

    with open(page_file) as f:
        content = f.read()

    # Check if there are ### subsections
    sections = []
    current_section = None
    current_profiles = []

    for line in content.split("\n"):
        if line.startswith("### "):
            if current_section and current_profiles:
                sections.append((current_section, current_profiles))
            current_section = line[4:].strip()
            current_profiles = []
        elif current_section and "| MII Elternprofil |" not in line and line.startswith("|"):
            # Table row - extract parent profile name (2nd column)
            cols = [c.strip() for c in line.split("|")]
            if len(cols) >= 4 and cols[2] and not cols[2].startswith("---"):
                parent = cols[2].strip()
                if parent:
                    current_profiles.append(parent)

    if current_section and current_profiles:
        sections.append((current_section, current_profiles))

    return sections


def _parse_basis_sections() -> list[tuple[str, list[str]]]:
    """Parse the 4 basis sub-module pages (person, diagnose, prozedur, fall)."""
    sections = []
    for sub, page_name in [
        ("Person", "modul-person.md"),
        ("Diagnose", "modul-diagnose.md"),
        ("Prozedur", "modul-prozedur.md"),
        ("Fall", "modul-fall.md"),
    ]:
        page_file = PAGECONTENT_DIR / page_name
        if not page_file.exists():
            continue
        with open(page_file) as f:
            content = f.read()

        parents = []
        for line in content.split("\n"):
            if "| MII Elternprofil |" in line or line.startswith("|---"):
                continue
            if line.startswith("|"):
                cols = [c.strip() for c in line.split("|")]
                if len(cols) >= 4 and cols[2] and not cols[2].startswith("---"):
                    parents.append(cols[2].strip())

        if parents:
            sections.append((sub, parents))

    return sections


# ---------------------------------------------------------------------------
# Per-element label resolution (shared between markdown + xlsx output)
# ---------------------------------------------------------------------------

def resolve_element_labels(
    elem: dict,
    resource_type: str,
    lm_lookup: dict | None,
    canonical: dict | None,
    element_comments: dict,
    pre_select_ids: set | None = None,
) -> dict | None:
    """Resolve all DE+EN labels for one MS element.

    Returns dict with keys: element, lm_label_de, lm_label_en, lm_desc,
    de_short, de_def, en_short, en_def, comment. Returns None if the element
    has no display path.
    """
    display_path = element_display_path(elem, resource_type)
    if not display_path:
        return None

    de_short = extract_translation(elem, "short", "de-DE")
    de_def = extract_translation(elem, "definition", "de-DE")
    en_short = extract_translation(elem, "short", "en-US")
    en_def = extract_translation(elem, "definition", "en-US")

    canon = lookup_canonical(canonical or {}, display_path)
    coding_label = {}
    if ".coding:" in display_path:
        cs = {}
        sys_url = get_coding_system(elem)
        if sys_url:
            cs = lookup_coding_system_labels(canonical or {}, sys_url)
        if not cs:
            slice_name = display_path.rsplit(":", 1)[-1]
            cs = coding_system_display_from_slice(slice_name)
        if cs:
            coding_label = {
                "de_short": cs.get("de_short", ""),
                "en_short": cs.get("en_short", ""),
                "de_def": f"Kodierung nach {cs.get('de_short', '')}.",
                "en_def": f"Coding in {cs.get('en_short', '')}.",
            }
            # For coding slices, coding_label is more reliable than upstream
            # Translation (upstream often has English-as-DE leaks like
            # "A reference to a code defined by SNOMED CT"). Override.
            de_short = coding_label["de_short"]
            de_def = coding_label["de_def"]
            en_short = coding_label["en_short"]
            en_def = coding_label["en_def"]

    if not de_short:
        de_short = canon.get("de_short") or elem.get("short", "")
    if not de_def:
        de_def = canon.get("de_def") or elem.get("definition", "")
    if not en_short:
        en_short = canon.get("en_short") or elem.get("short", "")
    if not en_def:
        en_def = canon.get("en_def") or elem.get("definition", "")

    lm_label_de = ""
    lm_label_en = ""
    lm_desc = ""
    if lm_lookup:
        lm_label_de = lookup_lm_label(lm_lookup, resource_type, display_path, elem, "de")
        lm_label_en = lookup_lm_label(lm_lookup, resource_type, display_path, elem, "en")
        lm_desc = lookup_lm_description(lm_lookup, resource_type, display_path, elem)

    elem_id = elem.get("id", "")
    pre_selected = bool(pre_select_ids and elem_id in pre_select_ids)

    return {
        "element": display_path,
        "lm_label_de": lm_label_de,
        "lm_label_en": lm_label_en,
        "lm_desc": lm_desc,
        "de_short": de_short,
        "de_def": de_def,
        "en_short": en_short,
        "en_def": en_def,
        "comment": element_comments.get(display_path, ""),
        "pre_selected": pre_selected,
    }


def collect_profile_data(
    profile_info: dict,
    profile_data: dict,
    resource_type: str,
    lm_lookup: dict | None = None,
    canonical: dict | None = None,
    field_config: dict | None = None,
) -> dict:
    """Collect all data for one profile (header info + per-element rows) for xlsx output."""
    ms_elements = get_ms_elements(profile_data)
    element_comments = profile_info.get("element_comments", {})
    pre_select_ids = get_pre_select_ids(profile_data, field_config or {})
    rows = []
    for elem in ms_elements:
        row = resolve_element_labels(elem, resource_type, lm_lookup, canonical, element_comments, pre_select_ids)
        if row:
            rows.append(row)
    return {
        "fdpg_name": profile_info["fdpg_name"],
        "fdpg_id": profile_info["fdpg_id"],
        "parent_name": profile_info["parent_name"],
        "title_de": profile_info.get("title_de", "") or _profile_display_name(profile_info["parent_name"]),
        "resource_type": resource_type,
        "description": profile_info.get("description", ""),
        "rows": rows,
    }


# ---------------------------------------------------------------------------
# Generate markdown for one profile
# ---------------------------------------------------------------------------

def generate_profile_section(
    profile_info: dict,
    profile_data: dict,
    resource_type: str,
    lm_lookup: dict[tuple[str, str], dict] | None = None,
    canonical: dict | None = None,
    field_config: dict | None = None,
) -> tuple[str, str]:
    """Generate German table and English details block for one profile.

    Returns (german_md, english_md).
    """
    ms_elements = get_ms_elements(profile_data)
    if not ms_elements:
        return "", ""

    pre_select_ids = get_pre_select_ids(profile_data, field_config or {})

    fdpg_name = profile_info["fdpg_name"]
    fdpg_id = profile_info["fdpg_id"]
    parent_name = profile_info["parent_name"]
    description = profile_info.get("description", "")
    title_de = profile_info.get("title_de", "")
    element_comments = profile_info.get("element_comments", {})

    german_lines = []
    english_lines = []

    display_name = title_de if title_de else _profile_display_name(parent_name)
    german_lines.append(
        f"#### {display_name} ({resource_type})\n"
    )
    german_lines.append(
        f"**FDPG Profil:** [{fdpg_name}](StructureDefinition-{fdpg_id}.html)"
        f" · **MII Elternprofil:** {parent_name}\n"
    )

    # Display profile description if available (skip generic placeholders)
    if description and not description.startswith("FDPG Profil -"):
        german_lines.append(f"{description}\n")

    # Determine whether to include the Kommentar column
    has_any_comment = bool(element_comments)

    # Determine whether any element has a logical model label
    has_lm_column = False
    if lm_lookup:
        for elem in ms_elements:
            dp = element_display_path(elem, resource_type)
            if dp and lookup_lm_label(lm_lookup, resource_type, dp, elem):
                has_lm_column = True
                break

    has_pre_select_column = bool(pre_select_ids)

    # Build table header (combinations: ±LM columns × ±Kommentar × ±Pre-Select)
    hdr_cols = ["Element"]
    if has_lm_column:
        hdr_cols += ["Konzept (LM)", "Beschreibung (LM)"]
    hdr_cols += ["Kurzbeschreibung (de)", "Definition (de)"]
    if has_pre_select_column:
        hdr_cols.append("Vorausgewählt")
    if has_any_comment:
        hdr_cols.append("Kommentar")
    german_lines.append("| " + " | ".join(hdr_cols) + " |")
    german_lines.append("|" + "|".join("---" for _ in hdr_cols) + "|")

    has_english = False
    english_rows = []

    for elem in ms_elements:
        display_path = element_display_path(elem, resource_type)
        if not display_path:
            continue

        # German: prefer translation → coding-system label (for *.coding:* slices)
        # → canonical element fallback → FHIR default (often English)
        de_short = extract_translation(elem, "short", "de-DE")
        de_def = extract_translation(elem, "definition", "de-DE")
        canon = lookup_canonical(canonical or {}, display_path)

        coding_label = {}
        if ".coding:" in display_path:
            cs = {}
            sys_url = get_coding_system(elem)
            if sys_url:
                cs = lookup_coding_system_labels(canonical or {}, sys_url)
            if not cs:
                slice_name = display_path.rsplit(":", 1)[-1]
                cs = coding_system_display_from_slice(slice_name)
            if cs:
                coding_label = {
                    "de_short": cs.get("de_short", ""),
                    "en_short": cs.get("en_short", ""),
                    "de_def": f"Kodierung nach {cs.get('de_short', '')}.",
                    "en_def": f"Coding in {cs.get('en_short', '')}.",
                }
                de_short = coding_label["de_short"]
                de_def = coding_label["de_def"]

        if not de_short:
            de_short = canon.get("de_short") or elem.get("short", "")
        if not de_def:
            de_def = canon.get("de_def") or elem.get("definition", "")

        # Look up element comment from FSH
        comment = element_comments.get(display_path, "")

        # Look up logical model concept + description
        lm_label = ""
        lm_desc = ""
        if has_lm_column and lm_lookup:
            lm_label = lookup_lm_label(lm_lookup, resource_type, display_path, elem)
            lm_desc = lookup_lm_description(lm_lookup, resource_type, display_path, elem)

        row_cols = [f"`{display_path}`"]
        if has_lm_column:
            row_cols += [md_escape(lm_label), md_escape(lm_desc)]
        row_cols += [md_escape(de_short), md_escape(de_def)]
        if has_pre_select_column:
            row_cols.append("✓" if elem.get("id") in pre_select_ids else "")
        if has_any_comment:
            row_cols.append(md_escape(comment))
        german_lines.append("| " + " | ".join(row_cols) + " |")

        # English: for coding-slices use coding_label override (upstream en-US
        # is often verbose FHIR-spec text); otherwise prefer upstream translation,
        # then canonical, then upstream short.
        en_short = extract_translation(elem, "short", "en-US")
        en_def = extract_translation(elem, "definition", "en-US")
        if coding_label:
            en_short = coding_label["en_short"]
            en_def = coding_label["en_def"]
        if not en_short:
            en_short = canon.get("en_short") or elem.get("short", "")
        if not en_def:
            en_def = canon.get("en_def") or elem.get("definition", "")
        if en_short or en_def:
            has_english = True
            english_rows.append(
                f"| `{display_path}` | {md_escape(en_short)} | {md_escape(en_def)} |"
            )

    english_md = ""
    if has_english:
        english_lines.append("<details>")
        english_lines.append(f"<summary>English translations - {display_name}</summary>\n")
        english_lines.append("| Element | Short (en) | Definition (en) |")
        english_lines.append("|---------|-----------|-----------------|")
        english_lines.extend(english_rows)
        english_lines.append("\n</details>\n")
        english_md = "\n".join(english_lines)

    return "\n".join(german_lines), english_md


def _profile_display_name(parent_name: str) -> str:
    """Create a human-readable display name from MII parent profile name.

    E.g. MII_PR_Labor_Laboruntersuchung -> Laboruntersuchung
    """
    parts = parent_name.split("_")
    # Skip MII, PR, and module prefix (first 3 parts)
    if len(parts) > 3:
        return " ".join(parts[3:])
    return parent_name


# ---------------------------------------------------------------------------
# Generate full page for one module
# ---------------------------------------------------------------------------

def generate_module_page(module_name: str, canonical: dict | None = None, field_config: dict | None = None) -> str:
    """Generate the full datenkatalog markdown page for a module."""
    cfg = MODULES[module_name]
    canonical = merge_module_labels(canonical or {}, module_name)
    package_dir = get_package_dir(cfg)

    if not package_dir.is_dir():
        print(f"  ERROR: Package directory not found: {package_dir}", file=sys.stderr)
        return ""

    profiles = parse_fsh_profiles(module_name)
    if not profiles:
        print(f"  WARNING: No FSH profiles found for {module_name}", file=sys.stderr)
        return ""

    # Build parent_name -> profile_info mapping
    parent_to_profile = {p["parent_name"]: p for p in profiles}

    # Load all parent profile JSONs
    parent_data: dict[str, dict] = {}
    for p in profiles:
        data = load_profile_json(package_dir, p["parent_name"])
        if data:
            parent_data[p["parent_name"]] = data
        else:
            print(f"  WARNING: Could not find parent profile {p['parent_name']} in {package_dir}", file=sys.stderr)

    # Load logical models and build label lookup
    lm_models = load_logical_models(module_name, package_dir)
    lm_lookup: dict[tuple[str, str], list[str]] | None = None
    if lm_models:
        lm_cfg = LOGICAL_MODEL_CONFIG.get(module_name, {})
        lm_lookup = build_lm_lookup(lm_models, lm_cfg.get("fhir_identities", []))

    # Get section groupings
    sections = parse_module_sections(module_name)

    # Build page
    lines = []
    lines.append(f"# Datenkatalog {cfg['title']}\n")

    if cfg.get("subtitle"):
        lines.append(f"*{cfg['subtitle']}*\n")

    lines.append(
        "Diese Seite listet alle MustSupport-Elemente der MII-Elternprofile mit "
        "deutschen und englischen Beschreibungen. "
        "Die Obligations werden auf der Seite [Obligations](obligations.html) beschrieben.\n"
    )

    lines.append(f"**Quellpaket:** [{cfg['package']}]({cfg['simplifier_url']})\n")

    english_blocks = []

    if sections:
        # Grouped layout
        seen_parents = set()
        for section_title, section_parents in sections:
            section_german = []
            section_english = []
            for parent_name in section_parents:
                if parent_name not in parent_to_profile:
                    continue
                if parent_name not in parent_data:
                    continue
                seen_parents.add(parent_name)
                pinfo = parent_to_profile[parent_name]
                pdata = parent_data[parent_name]
                resource_type = pdata.get("type", "Resource")
                german_md, english_md = generate_profile_section(pinfo, pdata, resource_type, lm_lookup, canonical, field_config)
                if german_md:
                    section_german.append(german_md)
                if english_md:
                    section_english.append(english_md)

            if section_german:
                lines.append(f"### {section_title}\n")
                for gmd in section_german:
                    lines.append(gmd)
                    lines.append("")

            english_blocks.extend(section_english)

        # Catch-all for profiles not in any section
        remaining = []
        for p in profiles:
            if p["parent_name"] not in seen_parents and p["parent_name"] in parent_data:
                remaining.append(p)

        if remaining:
            lines.append("### Weitere Profile\n")
            for pinfo in remaining:
                pdata = parent_data[pinfo["parent_name"]]
                resource_type = pdata.get("type", "Resource")
                german_md, english_md = generate_profile_section(pinfo, pdata, resource_type, lm_lookup, canonical, field_config)
                if german_md:
                    lines.append(german_md)
                    lines.append("")
                if english_md:
                    english_blocks.append(english_md)
    else:
        # Flat layout (no sub-sections)
        for pinfo in profiles:
            if pinfo["parent_name"] not in parent_data:
                continue
            pdata = parent_data[pinfo["parent_name"]]
            resource_type = pdata.get("type", "Resource")
            german_md, english_md = generate_profile_section(pinfo, pdata, resource_type, lm_lookup, canonical, field_config)
            if german_md:
                lines.append(german_md)
                lines.append("")
            if english_md:
                english_blocks.append(english_md)

    # Append all English blocks at the end
    if english_blocks:
        lines.append("---\n")
        lines.append("## English Translations\n")
        lines.extend(english_blocks)

    return "\n".join(lines) + "\n"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def collect_module_data(module_name: str, canonical: dict, field_config: dict | None = None) -> dict | None:
    """Collect all profile data for one module (for xlsx output)."""
    cfg = MODULES[module_name]
    canonical = merge_module_labels(canonical or {}, module_name)
    package_dir = get_package_dir(cfg)
    if not package_dir.is_dir():
        return None
    profiles = parse_fsh_profiles(module_name)
    if not profiles:
        return None

    lm_models = load_logical_models(module_name, package_dir)
    lm_lookup = None
    if lm_models:
        lm_cfg = LOGICAL_MODEL_CONFIG.get(module_name, {})
        lm_lookup = build_lm_lookup(lm_models, lm_cfg.get("fhir_identities", []))

    profile_data_list = []
    for p in profiles:
        pdata = load_profile_json(package_dir, p["parent_name"])
        if not pdata:
            continue
        resource_type = pdata.get("type", "Resource")
        collected = collect_profile_data(p, pdata, resource_type, lm_lookup, canonical, field_config)
        if collected["rows"]:
            profile_data_list.append(collected)

    return {
        "module": module_name,
        "title": cfg["title"],
        "subtitle": cfg.get("subtitle", ""),
        "package": cfg["package"],
        "version": cfg["version"],
        "simplifier_url": cfg["simplifier_url"],
        "profiles": profile_data_list,
    }


def write_xlsx(modules_data: list[dict], output_path: Path) -> None:
    """Write all module data to a single xlsx workbook (one sheet per module)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    headers = [
        "Profil (DE)", "FDPG ID", "MII Elternprofil", "Ressourcentyp",
        "Element",
        "Konzept LM (DE)", "Konzept LM (EN)", "Beschreibung LM",
        "Kurzbeschreibung (DE)", "Kurzbeschreibung (EN)",
        "Definition (DE)", "Definition (EN)",
        "Vorausgewählt",
        "Kommentar",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    body_align = Alignment(vertical="top", wrap_text=True)

    for module_info in modules_data:
        sheet_title = module_info["title"][:31]  # xlsx sheet name max 31 chars
        ws = wb.create_sheet(title=sheet_title)
        ws.append(headers)
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        for prof in module_info["profiles"]:
            for row in prof["rows"]:
                ws.append([
                    prof["title_de"],
                    prof["fdpg_id"],
                    prof["parent_name"],
                    prof["resource_type"],
                    row["element"],
                    row["lm_label_de"],
                    row["lm_label_en"],
                    row["lm_desc"],
                    row["de_short"],
                    row["en_short"],
                    row["de_def"],
                    row["en_def"],
                    "✓" if row.get("pre_selected") else "",
                    row["comment"],
                ])

        # Column widths (rough heuristic) and body alignment
        widths = [28, 32, 38, 14, 30, 22, 22, 40, 32, 32, 50, 50, 14, 32]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for row_cells in ws.iter_rows(min_row=2):
            for cell in row_cells:
                cell.alignment = body_align

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Generate FDPG Datenkatalog markdown + xlsx.")
    parser.add_argument("--no-md", action="store_true", help="Skip markdown generation")
    parser.add_argument("--no-xlsx", action="store_true", help="Skip xlsx generation")
    parser.add_argument("--xlsx-out", default="output/datenkatalog.xlsx",
                        help="xlsx output path (default: output/datenkatalog.xlsx)")
    args = parser.parse_args()

    print("Generating Datenkatalog...")
    canonical = load_canonical_labels()
    field_config = load_field_config()
    print(f"  Loaded {len(canonical.get('elements', {}))} canonical element labels, "
          f"{len(canonical.get('coding_systems', {}))} coding systems, "
          f"pre-select sets for {len(field_config)} FHIR base types")

    md_count = 0
    modules_data = []
    for module_name, cfg in MODULES.items():
        print(f"  Processing {module_name} ({cfg['title']})...")
        if not args.no_md:
            content = generate_module_page(module_name, canonical, field_config)
            if content:
                output_file = PAGECONTENT_DIR / f"datenkatalog-{module_name}.md"
                with open(output_file, "w") as f:
                    f.write(content)
                md_count += 1
                print(f"  -> {output_file.name}")
            else:
                print(f"  SKIPPED markdown for {module_name} (no content)")
        if not args.no_xlsx:
            data = collect_module_data(module_name, canonical, field_config)
            if data and data["profiles"]:
                modules_data.append(data)

    if not args.no_xlsx and modules_data:
        xlsx_path = PROJECT_ROOT / args.xlsx_out
        xlsx_path.parent.mkdir(parents=True, exist_ok=True)
        write_xlsx(modules_data, xlsx_path)
        print(f"\nWrote xlsx: {xlsx_path} ({len(modules_data)} sheets)")

    if not args.no_md:
        print(f"Generated {md_count} markdown pages.")


if __name__ == "__main__":
    main()
